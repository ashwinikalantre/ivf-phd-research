from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


path = Path("outputs/phd_ivf_lit_review/ivf_ai_literature_review_matrix_2021_2025.xlsx")
wb = load_workbook(path)
ws = wb["Literature Review Matrix"]

headers = [c.value for c in ws[1]]
existing_titles = {str(ws.cell(r, 2).value).strip().lower() for r in range(2, ws.max_row + 1)}

new_papers = [
    {
        "Title": "Comparative study of machine learning approaches integrated with genetic algorithm for IVF success prediction",
        "Authors": "Dehghan et al.",
        "Year": 2024,
        "Journal": "PLOS ONE",
        "Objective": "Compare ML algorithms with genetic-algorithm feature selection for IVF success prediction.",
        "Dataset": "Single medical facility in Tehran, Iran; IVF clinical dataset.",
        "Features Used": "Female age, AMH, endometrial thickness, sperm count, sperm morphology, follicle size, retrieved oocytes, oocyte quality, embryo quality and related variables.",
        "Method/Algorithm": "ANN, RPART, Random Forest, SVM and AdaBoost with Genetic Algorithm feature selection.",
        "Technology Used": "Machine learning, ensemble learning, genetic algorithm feature selection.",
        "Evaluation Metrics": "Accuracy; AdaBoost with GA 89.8%, Random Forest with GA 87.4%.",
        "Main Findings": "Feature selection improved all classifiers; AdaBoost and Random Forest performed strongly for IVF success prediction.",
        "Limitations": "Single medical facility in Tehran; limited demographic and healthcare-setting generalizability; needs external validation.",
        "Future Work": "External validation in other clinical settings and populations.",
        "Gap Found": "Single-center Study; Need External Validation; Limited Population Diversity; Need Clinical Decision Support",
        "Category": "IVF success prediction; Feature selection",
        "Relevance": "High",
        "Source URL": "https://journals.plos.org/plosone/article?id=10.1371/journal.pone.0310829",
        "Extraction Confidence": "High - student extraction plus publisher/PubMed verification",
    },
    {
        "Title": "Catalyzing IVF outcome prediction: exploring advanced machine learning paradigms for enhanced success rate prognostication",
        "Authors": "Sadegh-Zadeh et al.",
        "Year": 2024,
        "Journal": "Frontiers in Artificial Intelligence",
        "Objective": "Improve IVF success-rate prediction through advanced ML and gynecological expertise.",
        "Dataset": "Datasets from 2017-2018 and 2010-2016.",
        "Features Used": "Patient demographics, infertility factors, treatment protocols, maternal age, years of infertility, gonadotropin dosage, LH levels and ovarian response.",
        "Method/Algorithm": "Logistic Regression, Gaussian NB, SVM, MLP, KNN, Random Forest, AdaBoost, LogitBoost, RUSBoost and RSM.",
        "Technology Used": "Machine learning, ensemble learning, feature engineering.",
        "Evaluation Metrics": "Accuracy; LogitBoost reported at 96.35%.",
        "Main Findings": "Ensemble methods performed strongly; demographics, infertility factors and treatment protocols were important for IVF prediction.",
        "Limitations": "Retrospective design limits control of unknown confounders.",
        "Future Work": "Prospective studies with controlled variables and stronger clinical validation.",
        "Gap Found": "Retrospective Design; Need Prospective/RCT Validation; Need Clinical Decision Support",
        "Category": "IVF success prediction; Ensemble ML",
        "Relevance": "High",
        "Source URL": "https://www.frontiersin.org/journals/artificial-intelligence/articles/10.3389/frai.2024.1392611/full",
        "Extraction Confidence": "High - student extraction plus publisher/PMC verification",
    },
    {
        "Title": "Comparison of machine learning classification techniques to predict implantation success in an IVF treatment cycle",
        "Authors": "Yigit, Bener and Karabulut",
        "Year": 2022,
        "Journal": "Reproductive BioMedicine Online",
        "Objective": "Compare ML models for predicting implantation success in IVF cycles and variable importance.",
        "Dataset": "IVF treatment-cycle dataset; full dataset details require paper verification.",
        "Features Used": "Clinical IVF variables; maternal age identified as dominant predictor.",
        "Method/Algorithm": "Machine learning classification techniques.",
        "Technology Used": "ML classification.",
        "Evaluation Metrics": "Implantation prediction performance; exact values require full-paper extraction.",
        "Main Findings": "ML algorithms predicted implantation outcomes; maternal age was the most important predictor.",
        "Limitations": "Full limitations not captured in student extraction; likely requires validation and richer features.",
        "Future Work": "Full-paper reading needed before high-confidence use.",
        "Gap Found": "Need Full-text Verification; Need External Validation",
        "Category": "Implantation prediction",
        "Relevance": "Medium",
        "Source URL": "https://pubmed.ncbi.nlm.nih.gov/36088224/",
        "Extraction Confidence": "Medium - student extraction plus PubMed verification",
    },
    {
        "Title": "Predicting clinical pregnancy using clinical features and machine learning algorithms in in vitro fertilization",
        "Authors": "Wang et al.",
        "Year": 2022,
        "Journal": "PLOS ONE",
        "Objective": "Construct ML models for clinical pregnancy prediction in IVF using clinical features.",
        "Dataset": "IVF patients from 2007-2019.",
        "Features Used": "Twelve clinical variables; ovarian stimulation-related variable noted as important in extraction.",
        "Method/Algorithm": "Random Forest and Logistic Regression.",
        "Technology Used": "Machine learning prediction.",
        "Evaluation Metrics": "AUC: logistic regression 0.6766; Random Forest 0.7208.",
        "Main Findings": "Random Forest outperformed logistic regression for clinical pregnancy prediction.",
        "Limitations": "Full limitations not captured; temporal and external validation need verification.",
        "Future Work": "Full-paper extraction of limitations, dataset source and validation design.",
        "Gap Found": "Need Full-text Verification; Need External Validation; Need Clinical Decision Support",
        "Category": "Clinical pregnancy prediction",
        "Relevance": "Medium-High",
        "Source URL": "https://journals.plos.org/plosone/article?id=10.1371/journal.pone.0267554",
        "Extraction Confidence": "High - student extraction plus PLOS/PubMed verification",
    },
    {
        "Title": "Machine Learning Approach to Predict Clinical Pregnancy Potential in Women Undergoing IVF Program",
        "Authors": "Handayani et al.",
        "Year": 2022,
        "Journal": "Fertility & Reproduction",
        "Objective": "Use ML data mining to construct a clinical pregnancy prediction model for IVF.",
        "Dataset": "Large practical IVF clinical database; exact size/source require verification.",
        "Features Used": "Clinical IVF variables; no consensus feature protocol noted.",
        "Method/Algorithm": "Decision Tree and Random Forest.",
        "Technology Used": "Machine learning data mining.",
        "Evaluation Metrics": "Accuracy, precision, recall, F1 score; balanced accuracy below 0.7 noted.",
        "Main Findings": "DT and RF showed potential for clinical pregnancy prediction if sufficient data are available.",
        "Limitations": "Different datasets, predictor variables and metrics make studies difficult to compare; no consensus protocol for predictive variables.",
        "Future Work": "Standardize predictive variables and validation metrics across IVF ML studies.",
        "Gap Found": "Lack of Standardization; Need External Validation; Need Clinical Decision Support",
        "Category": "Clinical pregnancy prediction; Standardization",
        "Relevance": "High for gap argument",
        "Source URL": "https://www.worldscientific.com/doi/10.1142/S2661318222500098",
        "Extraction Confidence": "Medium - limitation is directly useful",
    },
    {
        "Title": "Artificial intelligence-driven precision treatment of reproductive medicine-related diseases: the optimal protocol choice for IVF-ET",
        "Authors": "Wen, Wu et al.",
        "Year": 2025,
        "Journal": "Journal of Advanced Research",
        "Objective": "Develop an AI system for personalized ovarian stimulation protocol choice in IVF-ET.",
        "Dataset": "17,791 patients undergoing OS and IVF/ICSI at Tongji Hospital, May 2015-May 2019.",
        "Features Used": "AMH, age, AFC, FSH, IVF-ET cycles and process indicators including progesterone, oocytes retrieved, estradiol and endometrial thickness.",
        "Method/Algorithm": "Adaptive ensemble/ACA-FI framework using RF, XGBoost and SOIL feature-selection approaches.",
        "Technology Used": "AI-based CDSS, ensemble learning.",
        "Evaluation Metrics": "Prediction of P, NOR, E2, endometrial thickness and pregnancy-related outcomes; exact metrics require paper verification.",
        "Main Findings": "AI can optimize ovarian stimulation protocol selection and provide clinically meaningful insights.",
        "Limitations": "Single large-scale center; possible selection bias from limited feature-selection algorithms; external validity concerns.",
        "Future Work": "Validate across multi-center and multi-population datasets; incorporate broader algorithms.",
        "Gap Found": "Single-center Study; Need External Validation; Need Clinical Decision Support; Personalized Recommendation Gap",
        "Category": "Protocol recommendation; CDSS",
        "Relevance": "Very High",
        "Source URL": "https://pubmed.ncbi.nlm.nih.gov/41139020/",
        "Extraction Confidence": "Medium-High - detailed student extraction",
    },
    {
        "Title": "Real-world use of an artificial intelligence-powered clinical decision support tool for ovarian stimulation",
        "Authors": "Bixby and Miller",
        "Year": 2025,
        "Journal": "F&S Reports",
        "Objective": "Evaluate real-world use of Stim Assist CDSS for FSH dosing and trigger timing during ovarian stimulation.",
        "Dataset": "Single IVF clinic real-world deployment cohort; exact size requires full-paper verification.",
        "Features Used": "Clinical ovarian stimulation variables used by Starting Dose Tool and Trigger Tool.",
        "Method/Algorithm": "AI-powered CDSS with Starting Dose and Trigger algorithms.",
        "Technology Used": "Stim Assist clinical decision support software.",
        "Evaluation Metrics": "Starting FSH dose, total FSH dose, laboratory outcomes.",
        "Main Findings": "AI guidance reduced starting and total FSH dosing without reducing laboratory outcomes.",
        "Limitations": "Single clinic; not randomized controlled trial; training data may not cover all scenarios.",
        "Future Work": "Prospective studies with larger populations and diverse clinics.",
        "Gap Found": "Single-center Study; Need Prospective/RCT Validation; Need Real-world Deployment Evidence; Need External Validation",
        "Category": "Real-world CDSS; Ovarian stimulation",
        "Relevance": "Very High",
        "Source URL": "https://www.fertstertreports.org/article/S2666-3341(25)00017-0/fulltext",
        "Extraction Confidence": "High - student extraction plus publisher verification",
    },
    {
        "Title": "A new decision-support tool in a multi-center randomized trial for personalized, optimized, and simplified fertility treatment in non-PCOS patients",
        "Authors": "Diwekar et al.",
        "Year": 2024,
        "Journal": "Reproduction and Fertility",
        "Objective": "Evaluate Opt-IVF CDSS for personalized gonadotropin dosing and reduced ultrasound monitoring in non-PCOS IVF patients.",
        "Dataset": "Multi-center randomized trial with 115 women aged 25-45; 55 intervention and 60 control.",
        "Features Used": "Follicle size distribution from day 1 and day 5, gonadotropin dosing, trigger timing and blastocyst outcomes.",
        "Method/Algorithm": "Physics-based model and optimal control theory decision-support tool.",
        "Technology Used": "Opt-IVF CDSS.",
        "Evaluation Metrics": "Total gonadotropin dose, repeated ultrasonograms, good-quality blastocysts and pregnancy rates.",
        "Main Findings": "Opt-IVF reduced hormone dosage and monitoring while improving good-quality blastocysts/pregnancy outcomes in reported trial.",
        "Limitations": "Non-PCOS population only; modest sample size; broader validation needed.",
        "Future Work": "Larger trials across broader patient groups including PCOS and diverse clinics.",
        "Gap Found": "Small Dataset; Limited Population Scope; Need External Validation; Need Clinical Decision Support",
        "Category": "Randomized CDSS; Personalization",
        "Relevance": "High",
        "Source URL": "https://pmc.ncbi.nlm.nih.gov/articles/PMC11466266/",
        "Extraction Confidence": "High - student extraction plus PMC verification",
    },
    {
        "Title": "Towards Automation in IVF: Pre-Clinical Validation of a Deep Learning-Based Embryo Grading System during PGT-A Cycles",
        "Authors": "Cimadomo et al.",
        "Year": 2023,
        "Journal": "Journal of Clinical Medicine",
        "Objective": "Validate a deep-learning embryo grading/ranking system during PGT-A cycles.",
        "Dataset": "Retrospective pre-clinical external validation including 3,604 blastocysts and 808 euploid transfers from 1,232 cycles.",
        "Features Used": "Time-lapse embryo videos, morphology, euploidy and live-birth associations.",
        "Method/Algorithm": "iDAScore v1.0 deep learning model using 3D CNN.",
        "Technology Used": "Deep learning embryo grading decision-support system.",
        "Evaluation Metrics": "AUC and associations with morphology, euploidy and live birth.",
        "Main Findings": "AI grading may objectify embryo evaluation and reduce manual annotation bias; performance comparable to embryologists for some outcomes.",
        "Limitations": "Retrospective design; needs prospective multicenter study; AI version not trained for all endpoints.",
        "Future Work": "Prospective multicenter validation and dataset enrichment.",
        "Gap Found": "Retrospective Design; Need External Validation; Need Prospective/RCT Validation; Black-box/Explainability Gap",
        "Category": "Embryo grading; Deep learning",
        "Relevance": "High",
        "Source URL": "https://www.mdpi.com/2077-0383/12/5/1806",
        "Extraction Confidence": "High - student extraction plus MDPI/PubMed verification",
    },
    {
        "Title": "An explainable ultrasound-based machine learning model for predicting reproductive outcomes after frozen embryo transfer",
        "Authors": "Xu et al.",
        "Year": 2025,
        "Journal": "Reproductive BioMedicine Online",
        "Objective": "Develop an explainable ultrasound radiomics plus clinical model for reproductive outcome prediction after FET.",
        "Dataset": "Prospective study of 787 infertile females undergoing FET; train n=550, test n=237.",
        "Features Used": "Endometrium and junctional-zone ultrasound radiomics, clinical data, radiomics score.",
        "Method/Algorithm": "Radiomics model, clinical logistic regression model, XGBoost fusion model, SHAP explanations.",
        "Technology Used": "Ultrasound radiomics, XGBoost, SHAP XAI.",
        "Evaluation Metrics": "AUC; fusion model training AUC 0.861.",
        "Main Findings": "Fusion model outperformed clinical and radiomics-only models; SHAP gave individual-level explanations.",
        "Limitations": "Single reproductive center; limited non-high-quality embryo transfers; FET only; manual ROI may introduce bias.",
        "Future Work": "Multicenter prospective validation; include larger non-high-quality embryo-transfer samples; automatic segmentation.",
        "Gap Found": "Single-center Study; Need External Validation; Need Multimodal Data; Manual/Subjective Workflow Bias",
        "Category": "FET prediction; Ultrasound radiomics; XAI",
        "Relevance": "Very High",
        "Source URL": "https://pubmed.ncbi.nlm.nih.gov/40199653/",
        "Extraction Confidence": "High - student extraction plus PubMed verification",
    },
    {
        "Title": "Interpretable, not black-box, artificial intelligence should be used for embryo selection",
        "Authors": "Afnan et al.",
        "Year": 2021,
        "Journal": "Human Reproduction Open",
        "Objective": "Argue for interpretable AI and rigorous evaluation before embryo-selection AI is ethically implemented.",
        "Dataset": "Commentary/review of embryo-selection AI literature.",
        "Features Used": "Not applicable.",
        "Method/Algorithm": "Ethical and methodological critique.",
        "Technology Used": "Interpretable AI/XAI for embryo selection.",
        "Evaluation Metrics": "Not applicable.",
        "Main Findings": "Existing AI can often separate good/poor embryos but not necessarily clinically difficult similar-quality embryos; black-box methods create trust and ethical risks.",
        "Limitations": "Commentary rather than empirical model; still highly relevant for gap framing.",
        "Future Work": "Use interpretable models, clinician-understandable explanations and RCT evaluation.",
        "Gap Found": "Black-box/Explainability Gap; Trustworthy AI Gap; Need Prospective/RCT Validation; Patient-centered XAI Gap",
        "Category": "XAI ethics; Embryo selection",
        "Relevance": "Very High for conceptual framing",
        "Source URL": "https://academic.oup.com/hropen/article/2021/4/hoab040/6415831",
        "Extraction Confidence": "Medium-High - strong extracted abstract/limitations",
    },
    {
        "Title": "Beyond black-box models: explainable AI for embryo ploidy prediction and patient-centric consultation",
        "Authors": "Luong et al.",
        "Year": 2024,
        "Journal": "Journal of Assisted Reproduction and Genetics",
        "Objective": "Use XAI to improve accuracy and transparency for embryo ploidy prediction using embryonic and clinical data.",
        "Dataset": "1,908 blastocyst embryos with ploidy status, morphokinetic features, morphology grades and 11 clinical variables.",
        "Features Used": "Morphokinetic features, morphology grades and clinical variables.",
        "Method/Algorithm": "RF, LDA, LR, SVM, AdaBoost and LightGBM with SHAP and LIME.",
        "Technology Used": "XAI, SHAP, LIME, embryo ploidy prediction.",
        "Evaluation Metrics": "Model performance across high-grade, low-grade and all-grade embryo datasets; exact values require full extraction.",
        "Main Findings": "XAI can improve transparency for ploidy prediction and patient-centric consultation.",
        "Limitations": "Full limitations not captured; validation/generalization need full-paper verification.",
        "Future Work": "Full-paper review required before using as core evidence.",
        "Gap Found": "Black-box/Explainability Gap; Patient-centered XAI Gap; Need External Validation",
        "Category": "Embryo ploidy; XAI",
        "Relevance": "High",
        "Source URL": "https://link.springer.com/article/10.1007/s10815-024-03178-7",
        "Extraction Confidence": "High - student extraction plus Springer/PubMed verification",
    },
]

next_no = max(ws.cell(r, 1).value for r in range(2, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, int)) + 1
for p in new_papers:
    if p["Title"].strip().lower() in existing_titles:
        continue
    row = []
    p = {"Paper No": next_no, **p}
    next_no += 1
    for h in headers:
        row.append(p.get(h, ""))
    ws.append(row)

thin = Side(style="thin", color="D9E2F3")
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 84

ws.auto_filter.ref = ws.dimensions
for table in ws.tables.values():
    table.ref = ws.dimensions

ws2 = wb["Gap Frequency"]
ws2.delete_rows(2, ws2.max_row)
gap_counter = Counter()
gap_papers = defaultdict(list)
for r in range(2, ws.max_row + 1):
    no = ws.cell(r, 1).value
    gaps = str(ws.cell(r, headers.index("Gap Found") + 1).value or "")
    for gap in [g.strip() for g in gaps.split(";") if g.strip()]:
        gap_counter[gap] += 1
        gap_papers[gap].append(str(no))
interpretation = {
    "Need External Validation": "Dominant methodological gap; essential before novelty or clinical usefulness can be claimed.",
    "Need Clinical Decision Support": "Prediction needs translation into clinician-facing recommendation workflows.",
    "Retrospective Design": "Limits causal and deployment claims.",
    "Single-center Study": "Restricts generalizability across clinics, populations and practice patterns.",
    "Trustworthy AI Gap": "Trust, transparency and responsible use must be designed and evaluated.",
    "Black-box/Explainability Gap": "Supports XAI as a core PhD direction, especially for embryo/patient counseling decisions.",
}
for gap, freq in gap_counter.most_common():
    ws2.append([gap, ", ".join(gap_papers[gap]), freq, interpretation.get(gap, "Repeated or emerging gap; assess after additional papers are added.")])
for table in ws2.tables.values():
    table.ref = ws2.dimensions
for row in ws2.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws3 = wb["Synthesis"]
for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
    if row[0].value == "Repeated Gap Pattern":
        row[1].value = "After adding the student's extracted papers, the strongest repeated gaps are external validation, clinical decision-support translation, single-center/limited-population evidence, retrospective design, and explainability/trust. The student papers strengthen the case for an XAI-enabled IVF CDSS rather than only an accuracy-focused prediction model."
    if row[0].value == "Current Supervisor Recommendation":
        row[1].value = "Proceed with a focused systematic review stage. Do not finalize the title yet, but the leading direction is now stronger: explainable and clinically validated personalized IVF decision support using multimodal clinical, embryology, ultrasound and lifestyle data, with special attention to Indian population validation."

wb.save(path)
print(path.resolve())
