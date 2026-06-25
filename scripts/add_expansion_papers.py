from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side


path = Path("outputs/phd_ivf_lit_review/ivf_ai_literature_review_matrix_2021_2025.xlsx")
wb = load_workbook(path)
ws = wb["Literature Review Matrix"]
headers = [c.value for c in ws[1]]
existing_titles = {str(ws.cell(r, 2).value).strip().lower() for r in range(2, ws.max_row + 1)}

new_papers = [
    {
        "Title": "Deep learning versus manual morphology-based embryo selection in IVF: a randomized, double-blind noninferiority trial",
        "Authors": "Illingworth et al.",
        "Year": 2024,
        "Journal": "Nature Medicine",
        "Objective": "Assess whether deep learning embryo selection is noninferior to standard morphology-based embryo selection.",
        "Dataset": "Multicenter randomized double-blind noninferiority trial across 14 IVF clinics in Australia and Europe; 1,066 patients, 533 per arm.",
        "Features Used": "Day-5 early-stage blastocyst images/time-lapse context used by iDAScore versus standard morphology assessment.",
        "Method/Algorithm": "Deep learning embryo selection using iDAScore compared with embryologist morphology-based selection.",
        "Technology Used": "Deep learning, embryo selection, randomized clinical trial.",
        "Evaluation Metrics": "Clinical pregnancy rate; risk difference; evaluation time.",
        "Main Findings": "Clinical pregnancy was 46.5% in the iDAScore arm versus 48.2% in morphology arm; noninferiority was not demonstrated, but evaluation time was reduced.",
        "Limitations": "Deep learning did not show noninferiority for clinical pregnancy; results do not prove superiority and reinforce the need for prospective clinical outcome validation.",
        "Future Work": "Further prospective trials and workflow studies to determine where AI adds value beyond speed and standardization.",
        "Gap Found": "Need Prospective/RCT Validation; Need Real-world Deployment Evidence; Trustworthy AI Gap; Clinical Utility Gap",
        "Category": "Embryo selection; RCT; Deep learning",
        "Relevance": "Very High",
        "Source URL": "https://www.nature.com/articles/s41591-024-03166-5",
        "Extraction Confidence": "High - Nature Medicine abstract/discussion and PubMed verification",
    },
    {
        "Title": "Artificial intelligence-enhanced in-vitro fertilization outcome prediction for the Indian subpopulation: integrating pre-treatment parameters and Bayesian-optimized ensemble techniques",
        "Authors": "Sengupta et al.",
        "Year": 2025,
        "Journal": "International Journal of Community Medicine and Public Health",
        "Objective": "Develop and evaluate an ML model for estimating cumulative live-birth probability in the first complete IVF cycle among Indian subpopulations.",
        "Dataset": "Retrospective cohort using IVF-ICSI data from Sir Ganga Ram Hospital, Delhi, India, from 1 January 2018 to 31 October 2022; original dataset 2,908 records and 79 variables.",
        "Features Used": "Pre-treatment clinical parameters; detailed final feature set requires full extraction.",
        "Method/Algorithm": "Bayesian-optimized ensemble model, reported as BoVe.",
        "Technology Used": "Machine learning, Bayesian optimization, cloud/web-app concept.",
        "Evaluation Metrics": "Reported model performance includes approximately 87% accuracy and 93% AUC in source summary.",
        "Main Findings": "Population-specific Indian model may support pretreatment counseling and personalized IVF planning.",
        "Limitations": "Single Indian center; retrospective design; needs external Indian multi-center validation and stronger peer-review/indexing verification.",
        "Future Work": "Validate across Indian IVF centers and assess clinical workflow integration.",
        "Gap Found": "No Indian Population; Single-center Study; Need External Validation; Retrospective Design; Need Clinical Decision Support",
        "Category": "Indian population; Live-birth prediction; Pretreatment prediction",
        "Relevance": "High for Indian validation angle",
        "Source URL": "https://www.ijcmph.com/index.php/ijcmph/article/view/13816",
        "Extraction Confidence": "Medium-High - journal page and PDF source reviewed",
    },
    {
        "Title": "Machine learning algorithms in constructing prediction models for assisted reproductive technology related live birth outcomes",
        "Authors": "Peng et al.",
        "Year": 2024,
        "Journal": "Scientific Reports",
        "Objective": "Construct and compare prediction models for ART-related live birth outcomes using clinical indicators.",
        "Dataset": "Large Chinese ART patient sample; exact cohort size requires full paper extraction.",
        "Features Used": "Maternal age, infertility duration, basal FSH, progressive sperm motility, E2, LH and progesterone on HCG day.",
        "Method/Algorithm": "Statistical models and multiple ML algorithms; logistic regression recommended for simplicity in source summary.",
        "Technology Used": "Machine learning and statistical prediction modeling.",
        "Evaluation Metrics": "Model comparison metrics reported in paper; exact values require full extraction.",
        "Main Findings": "Seven easily obtained demographic and clinical indicators were identified as promising live-birth predictors.",
        "Limitations": "Prediction-model literature has reporting, missing-data and validation limitations; full paper needed for study-specific limitations.",
        "Future Work": "Improve validation and reporting of ART live-birth prediction models.",
        "Gap Found": "Need Full-text Verification; Need External Validation; Lack of Standardization; Need Clinical Decision Support",
        "Category": "Live-birth prediction; ART; Clinical variables",
        "Relevance": "Medium-High",
        "Source URL": "https://www.nature.com/articles/s41598-024-83781-x",
        "Extraction Confidence": "Medium - abstract/source summary verified, full limitation extraction pending",
    },
    {
        "Title": "Machine learning prediction of clinical pregnancy in endometriosis patients undergoing fresh embryo transfer",
        "Authors": "Not fully extracted",
        "Year": 2025,
        "Journal": "BMC/PMC indexed article",
        "Objective": "Develop an ML-based predictive model for clinical pregnancy in endometriosis patients undergoing fresh embryo transfer.",
        "Dataset": "1,752 endometriosis patients undergoing IVF/ICSI with fresh embryo transfer from 2014 to 2024.",
        "Features Used": "Twenty-four clinical and embryonic characteristics.",
        "Method/Algorithm": "Naive Bayes, Logistic Regression, Random Forest, KNN, Neural Network and XGBoost with feature selection and cross-validation.",
        "Technology Used": "Machine learning, XGBoost, interpretable model analysis.",
        "Evaluation Metrics": "Clinical pregnancy prediction metrics reported; exact values require full extraction.",
        "Main Findings": "XGBoost effectively predicted clinical pregnancy and provided acceptable interpretability.",
        "Limitations": "Condition-specific retrospective cohort; broader population validation and deployment evidence still needed.",
        "Future Work": "Validate in external cohorts and assess clinical decision-support utility.",
        "Gap Found": "Disease-specific Generalization Gap; Retrospective Design; Need External Validation; Need Clinical Decision Support",
        "Category": "Endometriosis; Clinical pregnancy prediction; XGBoost",
        "Relevance": "High for personalization/subgroup theme",
        "Source URL": "https://pmc.ncbi.nlm.nih.gov/articles/PMC12406419/",
        "Extraction Confidence": "Medium-High - PMC abstract/source summary verified",
    },
    {
        "Title": "Artificial intelligence in assisted reproductive technology: separating the dream from reality",
        "Authors": "Cohen, Silvestri and Paredes",
        "Year": 2025,
        "Journal": "Reproductive BioMedicine Online",
        "Objective": "Critically review the evidence for AI in ART and separate realistic clinical value from overclaiming.",
        "Dataset": "Critical review of AI applications in ART/IVF.",
        "Features Used": "Not applicable; review covers embryo selection, IVF efficiency, standardization and outcome prediction.",
        "Method/Algorithm": "Critical narrative review.",
        "Technology Used": "AI in ART, clinical validation critique.",
        "Evaluation Metrics": "Not applicable.",
        "Main Findings": "AI has promise for efficiency and standardization, but many applications still lack rigorous validation.",
        "Limitations": "Review paper; useful for framing, not primary empirical evidence.",
        "Future Work": "Stronger validation, realistic clinical utility assessment and careful integration into ART practice.",
        "Gap Found": "Need External Validation; Clinical Utility Gap; Trustworthy AI Gap; Need Real-world Deployment Evidence",
        "Category": "Critical review; ART; Clinical utility",
        "Relevance": "High for thesis justification",
        "Source URL": "https://pubmed.ncbi.nlm.nih.gov/40287195/",
        "Extraction Confidence": "Medium-High - PubMed and journal metadata verified",
    },
    {
        "Title": "Development of an explainable machine learning model to predict live birth versus miscarriage among IVF-ET pregnancies",
        "Authors": "Not fully extracted",
        "Year": 2025,
        "Journal": "Journal of Medical Artificial Intelligence",
        "Objective": "Develop an interpretable ML model for live birth versus miscarriage prediction among patients with confirmed pregnancy after IVF-ET.",
        "Dataset": "IVF-ET pregnancies with confirmed pregnancy; exact cohort details require full extraction.",
        "Features Used": "Clinical variables including years of infertility and immune/inflammation markers in source figures.",
        "Method/Algorithm": "XGBoost explainable model with interpretable tree/feature analysis.",
        "Technology Used": "XGBoost, explainable machine learning.",
        "Evaluation Metrics": "AUC reported as 0.91 for live birth versus miscarriage prediction.",
        "Main Findings": "Interpretable ML may support post-transfer risk assessment in IVF-ET.",
        "Limitations": "Source explicitly highlights need for external validation in diverse populations.",
        "Future Work": "External validation across diverse populations and clinical integration studies.",
        "Gap Found": "Need External Validation; Need Clinical Decision Support; Trustworthy AI Gap; Need Full-text Verification",
        "Category": "Explainable AI; IVF-ET; Live birth versus miscarriage",
        "Relevance": "Medium-High",
        "Source URL": "https://jmai.amegroups.org/article/view/10224/html",
        "Extraction Confidence": "Medium - article page summary verified, full extraction pending",
    },
    {
        "Title": "A review of artificial intelligence applications in in vitro fertilization",
        "Authors": "Zhang, Liang and Chen",
        "Year": 2025,
        "Journal": "Journal of Assisted Reproduction and Genetics",
        "Objective": "Review recent AI applications across IVF, including embryo, gamete and clinical decision areas.",
        "Dataset": "Review paper.",
        "Features Used": "Not applicable; review covers AI applications across IVF workflow.",
        "Method/Algorithm": "Narrative review.",
        "Technology Used": "AI/ML/DL applications in IVF.",
        "Evaluation Metrics": "Not applicable.",
        "Main Findings": "AI may improve standardization, automation and IVF success, but translation requires careful validation.",
        "Limitations": "Review-level evidence; specific application claims depend on underlying primary studies.",
        "Future Work": "Broader validation and integration of AI into practical IVF workflows.",
        "Gap Found": "Need External Validation; Need Real-world Deployment Evidence; Lack of Standardization; Need Clinical Decision Support",
        "Category": "Review; IVF AI applications",
        "Relevance": "High for background and trend analysis",
        "Source URL": "https://pubmed.ncbi.nlm.nih.gov/39400647/",
        "Extraction Confidence": "High - PubMed/PMC metadata verified",
    },
    {
        "Title": "Global trends in the use of artificial intelligence in reproductive medicine: insights from surveys of international fertility specialists",
        "Authors": "Shoham et al.",
        "Year": 2025,
        "Journal": "Journal of IVF-Worldwide",
        "Objective": "Assess global AI adoption and attitudes in reproductive medicine using surveys of fertility specialists.",
        "Dataset": "Survey data from 383 eligible respondents in 2022 and 171 primarily fertility-specialist respondents in 2025.",
        "Features Used": "Survey responses on AI use, investment likelihood, perceived patient outcome impact and research/education uses.",
        "Method/Algorithm": "Survey analysis.",
        "Technology Used": "AI adoption and clinical workflow survey.",
        "Evaluation Metrics": "Survey proportions; AI usage trend from 2022 to 2025.",
        "Main Findings": "AI use in IVF/reproductive medicine increased substantially; embryo selection remained a common application, but many respondents still reported insufficient outcome data.",
        "Limitations": "Survey evidence; possible respondent and sampling bias; does not validate model performance.",
        "Future Work": "Link adoption trends with clinical evidence, training, governance and outcome evaluation.",
        "Gap Found": "Need Real-world Deployment Evidence; Trustworthy AI Gap; Clinical Utility Gap; Need Clinician Usability Evaluation",
        "Category": "AI adoption; Survey; Clinical workflow",
        "Relevance": "Medium-High for deployment context",
        "Source URL": "https://jivfww.scholasticahq.com/article/140673-global-trends-in-the-use-of-artificial-intelligence-ai-in-reproductive-medicine-insights-from-surveys-of-international-fertility-specialists",
        "Extraction Confidence": "High - journal page/PDF verified",
    },
]

next_no = max(ws.cell(r, 1).value for r in range(2, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, int)) + 1
for paper in new_papers:
    if paper["Title"].strip().lower() in existing_titles:
        continue
    paper = {"Paper No": next_no, **paper}
    next_no += 1
    ws.append([paper.get(h, "") for h in headers])

thin = Side(style="thin", color="D9E2F3")
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 84
ws.auto_filter.ref = ws.dimensions
for table in ws.tables.values():
    table.ref = ws.dimensions

ws2 = wb["Gap Frequency"]
ws2.delete_rows(2, ws2.max_row)
gap_counter = Counter()
gap_papers = defaultdict(list)
for r in range(2, ws.max_row + 1):
    no = ws.cell(r, 1).value
    gaps = str(ws.cell(r, headers.index("Gap Found") + 1).value or "")
    for gap in [g.strip() for g in gaps.split(";") if g.strip()]:
        gap_counter[gap] += 1
        gap_papers[gap].append(str(no))
for gap, freq in gap_counter.most_common():
    ws2.append([gap, ", ".join(gap_papers[gap]), freq, "Updated after Phase 2 expansion."])
for table in ws2.tables.values():
    table.ref = ws2.dimensions

wb.save(path)
print(path.resolve())
