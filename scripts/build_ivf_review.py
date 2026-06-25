from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference


OUT = Path("outputs/phd_ivf_lit_review/ivf_ai_literature_review_matrix_2021_2025.xlsx")


papers = [
    {
        "Paper No": 1,
        "Title": "Explainable artificial intelligence to identify follicles that optimize clinical outcomes during assisted conception",
        "Authors": "Hanassab et al.",
        "Year": 2025,
        "Journal": "Nature Communications",
        "Objective": "Use XAI/ML to identify follicle-size ranges associated with mature oocytes and downstream IVF outcomes.",
        "Dataset": "19,082 treatment-naive female patients from 11 European IVF centers.",
        "Features Used": "Individual follicle counts by size on day of trigger and prior days; age, BMI, stimulation days, trigger type, protocol, estradiol in secondary analyses.",
        "Method/Algorithm": "Histogram-based gradient boosting regression tree; permutation importance; SHAP; logistic regression for live birth associations.",
        "Technology Used": "Machine learning, XAI, SHAP/permutation importance.",
        "Evaluation Metrics": "MAE, MedAE, R-squared; odds ratios for live birth associations.",
        "Main Findings": "Follicles around 13-18 mm contributed most to mature oocyte retrieval; larger mean follicle size and >18 mm follicles were associated with progesterone elevation and lower fresh-transfer live birth.",
        "Limitations": "Needs prospective randomized validation before trigger-decision rules can be adopted clinically.",
        "Future Work": "Prospective trial comparing full follicle-cohort trigger strategy with threshold-based practice.",
        "Gap Found": "Need Prospective/RCT Validation; Need Clinical Decision Support; External/Indian Validation Gap",
        "Category": "XAI; Personalization; Ovarian stimulation",
        "Relevance": "High",
        "Source URL": "https://www.nature.com/articles/s41467-024-55301-y",
        "Extraction Confidence": "High - full page abstract/discussion/limitations available",
    },
    {
        "Paper No": 2,
        "Title": "Optimizing oocyte yield utilizing a machine learning model for dose and trigger decisions, a multi-center, prospective study",
        "Authors": "Scientific Reports authors",
        "Year": 2024,
        "Journal": "Scientific Reports",
        "Objective": "Evaluate clinical outcomes when clinicians use an AI platform for starting FSH dose and trigger-timing decisions.",
        "Dataset": "Prospective study with historical propensity-matched control; two US clinics, four physicians.",
        "Features Used": "Clinical ovarian stimulation inputs used by dose and trigger tools; exact feature set not fully extracted in this first pass.",
        "Method/Algorithm": "Commercial/clinical ML decision-support tools for starting dose and trigger timing.",
        "Technology Used": "AI clinical decision support platform for ovarian stimulation.",
        "Evaluation Metrics": "Oocyte yield and downstream embryology/clinical outcomes; statistical significance against matched controls.",
        "Main Findings": "Prospective clinical use was feasible; study designed as an early clinical evaluation rather than definitive outcome trial.",
        "Limitations": "Not randomized; only two clinics and four physicians; limited sample size; could not isolate starting-dose versus trigger-tool effects.",
        "Future Work": "Larger prospective multi-site studies with more physicians; isolate component algorithms; assess pregnancy and live-birth outcomes.",
        "Gap Found": "Need Prospective/RCT Validation; Small Dataset; Limited Multi-center Validation; Need Real-world Deployment Evidence",
        "Category": "Clinical decision support; Personalization; Deployment",
        "Relevance": "High",
        "Source URL": "https://www.nature.com/articles/s41598-024-69165-1",
        "Extraction Confidence": "High - full page abstract/discussion/limitations available",
    },
    {
        "Paper No": 3,
        "Title": "Multimodal intelligent prediction model for in vitro fertilization",
        "Authors": "npj Digital Medicine authors",
        "Year": 2025,
        "Journal": "npj Digital Medicine",
        "Objective": "Develop a multimodal IVF prediction model integrating embryo time-lapse imagery and clinical factors.",
        "Dataset": "Chinese IVF data; day-3 transferred embryos only.",
        "Features Used": "TLS embryo image/video-derived features plus maternal/clinical variables including age, AMH, gonadotropins, endometrial thickness, BMI and related health indicators.",
        "Method/Algorithm": "Multimodal deep learning model for post-transfer outcome prediction.",
        "Technology Used": "Deep learning, multimodal AI, clinical-image fusion.",
        "Evaluation Metrics": "Predictive performance metrics reported in paper; exact values not extracted in this first pass.",
        "Main Findings": "Multidimensional clinical data can improve personalized IVF outcome prediction beyond morphology alone.",
        "Limitations": "Geographically restricted to China; transferred day-3 embryos only; not exposed to non-transferred embryos; retrospective design.",
        "Future Work": "External international multi-center validation; include non-transferred embryos; prospective validation; extend toward pre-transfer preference recommendation.",
        "Gap Found": "Need External Validation; Limited Population Diversity; No Pre-transfer Recommendation; Retrospective Design",
        "Category": "Multimodal AI; Personalized prediction; Embryo selection",
        "Relevance": "High",
        "Source URL": "https://www.nature.com/articles/s41746-025-02331-5",
        "Extraction Confidence": "High - full page limitations available",
    },
    {
        "Paper No": 4,
        "Title": "Deep learning classification integrating embryo images with associated clinical information from ART cycles",
        "Authors": "Scientific Reports authors",
        "Year": 2025,
        "Journal": "Scientific Reports",
        "Objective": "Test whether embryo-image models improve when combined with patient clinical information.",
        "Dataset": "Embryo images and associated ART-cycle clinical data; transferred embryos only; limited database.",
        "Features Used": "Blastocyst/embryo images plus clinical metadata.",
        "Method/Algorithm": "Deep learning image model with clinical-data fusion.",
        "Technology Used": "CNN/deep learning, image-clinical fusion.",
        "Evaluation Metrics": "Classification performance metrics reported in paper; exact values not extracted in this first pass.",
        "Main Findings": "Fusion of images and clinical information produced more informed predictions than single-modality approaches.",
        "Limitations": "Class imbalance; only successfully transferred embryos; no discarded/frozen embryos; limited database.",
        "Future Work": "Include time-lapse videos and data from transferred, discarded, and frozen embryos; improve sample diversity.",
        "Gap Found": "Small Dataset; Selection Bias; Need Multimodal Data; No Pre-transfer Recommendation",
        "Category": "Multimodal AI; Embryo selection",
        "Relevance": "High",
        "Source URL": "https://www.nature.com/articles/s41598-025-02076-x",
        "Extraction Confidence": "High - full page limitations available",
    },
    {
        "Paper No": 5,
        "Title": "Machine learning center-specific models show improved IVF live birth predictions over US national registry-based model",
        "Authors": "Yao et al.",
        "Year": 2025,
        "Journal": "Nature Communications",
        "Objective": "Compare center-specific ML live-birth prediction with a national registry-based model.",
        "Dataset": "4,635 first-IVF-cycle patients from six centers.",
        "Features Used": "First-cycle patient and center-level IVF variables; exact feature list not extracted in this first pass.",
        "Method/Algorithm": "Center-specific machine learning live-birth prediction models.",
        "Technology Used": "Machine learning predictive analytics.",
        "Evaluation Metrics": "Precision-recall AUC and F1 score.",
        "Main Findings": "Center-specific models improved live-birth prediction compared with US registry-based prediction.",
        "Limitations": "US-center data; center-specific calibration may limit direct transfer to other clinics/populations.",
        "Future Work": "Validate and recalibrate across additional clinics and populations.",
        "Gap Found": "Limited Population Diversity; No Indian Population; Generalization Gap; Need External Validation",
        "Category": "Live-birth prediction; Generalization",
        "Relevance": "High",
        "Source URL": "https://www.nature.com/articles/s41467-025-58744-z",
        "Extraction Confidence": "Medium - publisher/search abstract used",
    },
    {
        "Paper No": 6,
        "Title": "Predictive modeling of pregnancy outcomes utilizing multiple machine learning techniques for IVF-ET",
        "Authors": "Bai et al.",
        "Year": 2025,
        "Journal": "BMC Pregnancy and Childbirth",
        "Objective": "Identify factors influencing pregnancy outcomes and compare ML models for IVF-ET prediction.",
        "Dataset": "2,625 women undergoing fresh IVF-ET cycles from 2016-2022 at Inner Mongolia Medical University affiliated hospital.",
        "Features Used": "Clinical IVF-ET variables after preprocessing; exact feature list not fully extracted.",
        "Method/Algorithm": "Multiple ML models; 80:20 train-test split.",
        "Technology Used": "Machine learning predictive modeling.",
        "Evaluation Metrics": "Accuracy-oriented model comparison; exact metrics not extracted in this first pass.",
        "Main Findings": "ML models can support clinical pregnancy outcome prediction and identify influential variables.",
        "Limitations": "Single-center retrospective dataset; limited evidence of external validation.",
        "Future Work": "External validation and clinical implementation assessment.",
        "Gap Found": "Single-center Study; Need External Validation; Retrospective Design; Need Clinical Decision Support",
        "Category": "Pregnancy prediction; ML",
        "Relevance": "High",
        "Source URL": "https://link.springer.com/article/10.1186/s12884-025-07433-2",
        "Extraction Confidence": "Medium - abstract/search source used",
    },
    {
        "Paper No": 7,
        "Title": "Construction and evaluation of machine learning-based prediction model for live birth following fresh embryo transfer in IVF/ICSI patients with PCOS",
        "Authors": "Zhu et al.",
        "Year": 2025,
        "Journal": "Journal of Ovarian Research",
        "Objective": "Predict live birth after fresh embryo transfer among PCOS patients using ML algorithms.",
        "Dataset": "1,062 fresh embryo transfer cycles in PCOS patients; 466 live births.",
        "Features Used": "Embryo transfer count, embryo type, maternal age, infertility duration, BMI, testosterone, progesterone on hCG day and related clinical variables.",
        "Method/Algorithm": "Multiple ML algorithms; XGBoost best; SHAP for predictors.",
        "Technology Used": "Machine learning, XGBoost, SHAP.",
        "Evaluation Metrics": "Comparative predictive performance; exact values not extracted in this first pass.",
        "Main Findings": "XGBoost showed superior predictive capacity and identified clinically relevant predictors.",
        "Limitations": "Condition-specific cohort; likely retrospective; external validation not evident in abstract-level extraction.",
        "Future Work": "Validate in broader PCOS populations and translate predictors into intervention-oriented decision support.",
        "Gap Found": "Need External Validation; Disease-specific Generalization Gap; Need Clinical Decision Support",
        "Category": "PCOS; Live-birth prediction; XAI",
        "Relevance": "Medium-High",
        "Source URL": "https://pmc.ncbi.nlm.nih.gov/articles/PMC11969817/",
        "Extraction Confidence": "Medium - abstract/search snippets used",
    },
    {
        "Paper No": 8,
        "Title": "Predictive models for live birth outcomes following fresh embryo transfer",
        "Authors": "Translational Medicine authors",
        "Year": 2025,
        "Journal": "Journal of Translational Medicine",
        "Objective": "Develop ML prediction models for live birth before embryo transfer.",
        "Dataset": "Fresh embryo-transfer cohort; exact size not extracted in first pass.",
        "Features Used": "Clinical and embryo-transfer variables; exact feature list not extracted.",
        "Method/Algorithm": "Machine learning prediction models.",
        "Technology Used": "ML predictive medicine.",
        "Evaluation Metrics": "Predictive performance metrics reported in paper; exact values not extracted.",
        "Main Findings": "ML prediction before embryo transfer may improve decision-making and patient counseling.",
        "Limitations": "External validation and deployment evidence not established from abstract-level extraction.",
        "Future Work": "Prospective and external validation; integration into counseling workflow.",
        "Gap Found": "Need External Validation; Need Clinical Decision Support; Need Real-world Deployment Evidence",
        "Category": "Live-birth prediction; Counseling",
        "Relevance": "Medium-High",
        "Source URL": "https://link.springer.com/article/10.1186/s12967-025-07045-6",
        "Extraction Confidence": "Medium - publisher abstract/search source used",
    },
    {
        "Paper No": 9,
        "Title": "A retrospective analysis of 48,514 IVF cycles and an evaluation of machine learning models for live birth prediction",
        "Authors": "Frontiers in Endocrinology authors",
        "Year": 2025,
        "Journal": "Frontiers in Endocrinology",
        "Objective": "Compare CNN and traditional ML models for live-birth prediction and assess feasibility in resource-limited settings.",
        "Dataset": "48,514 IVF cycles from EMR data.",
        "Features Used": "Electronic medical record variables from IVF cycles; exact variables not fully extracted.",
        "Method/Algorithm": "CNN, Random Forest and other traditional ML models.",
        "Technology Used": "Machine learning, CNN, EMR analytics.",
        "Evaluation Metrics": "AUC, F1 score, recall.",
        "Main Findings": "Random Forest and custom CNN performed strongly; CNN showed very high recall and F1 in reported summary.",
        "Limitations": "Retrospective EMR design; potential site/data-quality dependence; external validation not evident in abstract-level extraction.",
        "Future Work": "Validate prospectively and in resource-limited clinical settings.",
        "Gap Found": "Retrospective Design; Need External Validation; Need Real-world Deployment Evidence",
        "Category": "Live-birth prediction; EMR analytics",
        "Relevance": "High",
        "Source URL": "https://www.frontiersin.org/journals/endocrinology/articles/10.3389/fendo.2025.1556681/full",
        "Extraction Confidence": "Medium - abstract/search source used",
    },
    {
        "Paper No": 10,
        "Title": "Optimizing predictive features using machine learning for early miscarriage risk following single vitrified-warmed blastocyst transfer",
        "Authors": "Liu et al.",
        "Year": 2025,
        "Journal": "Frontiers in Endocrinology",
        "Objective": "Predict early miscarriage risk after single vitrified-warmed blastocyst transfer.",
        "Dataset": "Dual-center retrospective analysis of 1,664 SVBT cycles, including 308 early miscarriage cases.",
        "Features Used": "Maternal age, paternal age, endometrial thickness, blastocyst quality, ovarian stimulation parameters and related clinical variables.",
        "Method/Algorithm": "Logistic Regression, Random Forest, Gradient Boosting, Voting Classifier.",
        "Technology Used": "Ensemble machine learning.",
        "Evaluation Metrics": "AUC, accuracy, precision, recall, F1 score, specificity.",
        "Main Findings": "Voting Classifier achieved highest reported AUC, accuracy, precision and specificity, outperforming logistic regression.",
        "Limitations": "Dual-center retrospective study; needs broader external validation and clinical workflow testing.",
        "Future Work": "External prospective validation and individualized ART counseling integration.",
        "Gap Found": "Retrospective Design; Limited Multi-center Validation; Need Clinical Decision Support",
        "Category": "Miscarriage prediction; Personalized care",
        "Relevance": "Medium",
        "Source URL": "https://www.frontiersin.org/journals/endocrinology/articles/10.3389/fendo.2025.1557667/full",
        "Extraction Confidence": "High - full abstract/search result available",
    },
    {
        "Paper No": 11,
        "Title": "Machine learning-based prediction of IVF outcomes: the central role of female preprocedural factors",
        "Authors": "Bereczki et al.",
        "Year": 2025,
        "Journal": "Biomedicines",
        "Objective": "Develop and validate per-cycle IVF success prediction using preprocedural clinical factors.",
        "Dataset": "IVF cycles; exact size not extracted in first pass.",
        "Features Used": "Female preprocedural factors and baseline clinical variables.",
        "Method/Algorithm": "Machine learning prediction models.",
        "Technology Used": "ML predictive analytics.",
        "Evaluation Metrics": "Model performance metrics reported in paper; exact values not extracted.",
        "Main Findings": "Female preprocedural factors appear central for IVF outcome prediction.",
        "Limitations": "Detailed limitations require full-paper extraction; likely external validation and deployment remain open.",
        "Future Work": "Full extraction needed; likely validation across clinics and populations.",
        "Gap Found": "Need Full-text Verification; Need External Validation; Need Clinical Decision Support",
        "Category": "Preprocedural prediction; ML",
        "Relevance": "Medium",
        "Source URL": "https://pmc.ncbi.nlm.nih.gov/articles/PMC12649832/",
        "Extraction Confidence": "Low-Medium - abstract/search source only",
    },
    {
        "Paper No": 12,
        "Title": "The prospect of artificial intelligence to personalize assisted reproduction",
        "Authors": "Nature Medicine/npj Digital Medicine review authors",
        "Year": 2024,
        "Journal": "npj Digital Medicine",
        "Objective": "Review current and future AI implementations for personalized ART/IVF.",
        "Dataset": "Narrative/review synthesis of ART AI studies.",
        "Features Used": "Not applicable; review covers sperm, oocyte, embryo, clinical and treatment-planning data.",
        "Method/Algorithm": "Review of AI applications and translational barriers.",
        "Technology Used": "AI/ML/DL across ART.",
        "Evaluation Metrics": "Not applicable.",
        "Main Findings": "AI can support personalization in ART, but applications remain fragmented and translational evidence is limited.",
        "Limitations": "Highlights lack of standardization, validation, and clinical integration across AI-IVF studies.",
        "Future Work": "Rigorous validation, standardized datasets, clinical integration, and attention to trust/ethics.",
        "Gap Found": "Lack of Standardization; Need External Validation; Trustworthy AI Gap; Need Clinical Decision Support",
        "Category": "Review; Personalization; Trustworthy AI",
        "Relevance": "High",
        "Source URL": "https://www.nature.com/articles/s41746-024-01006-x",
        "Extraction Confidence": "Medium - review abstract/search source used",
    },
    {
        "Paper No": 13,
        "Title": "Current progress and open challenges for applying artificial intelligence in IVF",
        "Authors": "Patterns review authors",
        "Year": 2025,
        "Journal": "Patterns",
        "Objective": "Summarize AI-IVF applications by data modality and identify open challenges.",
        "Dataset": "Review of structured clinical, image, video, omics and multimodal AI-IVF studies.",
        "Features Used": "Not applicable; review stratifies by modality.",
        "Method/Algorithm": "Review/synthesis.",
        "Technology Used": "AI/ML/DL and multimodal AI.",
        "Evaluation Metrics": "Not applicable.",
        "Main Findings": "AI-IVF is promising but depends on validation, ethical safeguards and interdisciplinary clinical translation.",
        "Limitations": "Review-level limitation; primary evidence remains heterogeneous and validation-limited.",
        "Future Work": "Rigorous validation, ethical safeguards, interdisciplinary deployment, and better multimodal datasets.",
        "Gap Found": "Need External Validation; Trustworthy AI Gap; Need Multimodal Data; Need Real-world Deployment Evidence",
        "Category": "Review; Open challenges; Trustworthy AI",
        "Relevance": "High",
        "Source URL": "https://pmc.ncbi.nlm.nih.gov/articles/PMC12664965/",
        "Extraction Confidence": "Medium - abstract/search source used",
    },
    {
        "Paper No": 14,
        "Title": "Patient-centred explainability in IVF outcome prediction",
        "Authors": "Sivaprasad et al.",
        "Year": 2025,
        "Journal": "arXiv preprint",
        "Objective": "Evaluate understandability and trust in an IVF outcome prediction interface from patient feedback, survey and interviews.",
        "Dataset": "Four years of anonymous patient feedback plus user survey and interviews.",
        "Features Used": "Patient feedback themes, trust/understandability responses; model feature-space explanations.",
        "Method/Algorithm": "Human-centered XAI/interface evaluation.",
        "Technology Used": "Patient-facing explainable AI interface; proposed dialogue-based explanation.",
        "Evaluation Metrics": "Trust and understandability measures; qualitative feedback.",
        "Main Findings": "Patients need explainability beyond model feature space; concerns include data shifts and model exclusions.",
        "Limitations": "Preprint; not a clinical prediction validation study; generalizability needs peer-reviewed confirmation.",
        "Future Work": "Dialogue-based personalized explanations and patient-centered XAI methods.",
        "Gap Found": "Patient-centered XAI Gap; Trustworthy AI Gap; Need Personalized Explanation",
        "Category": "XAI; Trust; Patient interface",
        "Relevance": "Medium-High",
        "Source URL": "https://arxiv.org/abs/2506.18760",
        "Extraction Confidence": "Medium - preprint abstract used",
    },
    {
        "Paper No": 15,
        "Title": "AI-based live birth prediction in IVF cycles: a systematic review without meta-analysis of model performance and validation",
        "Authors": "Systematic review authors",
        "Year": 2026,
        "Journal": "Middle East Fertility Society Journal",
        "Objective": "Systematically review AI-based live-birth prediction models in IVF cycles.",
        "Dataset": "Systematic review of live-birth prediction studies; included because it directly synthesizes 2021-2025 evidence though published after target window.",
        "Features Used": "Not applicable; review of model inputs across studies.",
        "Method/Algorithm": "Systematic review without meta-analysis.",
        "Technology Used": "AI/ML model validation synthesis.",
        "Evaluation Metrics": "Model performance and validation practices across studies.",
        "Main Findings": "Evidence favors multimodal prediction but remains dominated by retrospective cohorts.",
        "Limitations": "Selection bias and internal-validation overfitting are recurring concerns across the evidence base.",
        "Future Work": "Stronger external/prospective validation and clearer model performance reporting.",
        "Gap Found": "Retrospective Design; Selection Bias; Internal Validation Overfitting; Need External Validation",
        "Category": "Systematic review; Live-birth prediction",
        "Relevance": "High as synthesis, outside year window",
        "Source URL": "https://link.springer.com/article/10.1186/s43043-026-00334-0",
        "Extraction Confidence": "Medium - search/publisher summary used",
    },
]


deliverables = [
    ("Research Trend Analysis", "2021-2025 work concentrates on embryo assessment, live-birth/pregnancy prediction, ovarian stimulation, and multimodal clinical-image fusion. 2024-2025 papers show a shift from accuracy-only models toward XAI, SHAP, decision support and prospective/clinical validation, but deployment evidence is still thin."),
    ("Repeated Gap Pattern", "The strongest repeated gaps are external/prospective validation, clinical decision-support translation, population diversity including Indian validation, multimodal longitudinal data, and trustworthy/patient-centered explainability."),
    ("Research Problem Statement", "IVF clinics generate rich clinical, embryological and lifestyle-relevant data, yet current AI models are often retrospective, population-specific, weakly validated outside their source clinic, and insufficiently explainable for clinician-patient shared decisions. This limits trustworthy personalized treatment recommendation and clinical decision support, especially for Indian IVF populations."),
    ("RQ1", "Which clinical, embryological and lifestyle variables most consistently predict IVF pregnancy/live-birth outcomes in recent AI-IVF literature?"),
    ("RQ2", "Can an explainable multimodal ML model improve individualized IVF outcome prediction compared with conventional or single-modality models?"),
    ("RQ3", "How can XAI outputs be translated into clinician-usable and patient-understandable recommendations without overstating causality?"),
    ("RQ4", "Does the model generalize across clinics or populations, and what recalibration is required for Indian IVF cohorts?"),
    ("Objective 1", "Build a systematic literature matrix for AI/ML/XAI in IVF from 2021-2025 and quantify repeated research gaps."),
    ("Objective 2", "Design an explainable prediction and recommendation framework using clinical, treatment-cycle, embryology and lifestyle variables where available."),
    ("Objective 3", "Evaluate predictive performance, calibration, explainability, fairness/subgroup behavior and clinical usability."),
    ("Objective 4", "Propose a clinical decision-support workflow for personalized IVF counseling and treatment planning."),
    ("Conceptual Framework", "Inputs: demographics, infertility history, hormones/AMH, ovarian stimulation, follicle profile, embryo/image features, endometrium and lifestyle factors. Processing: preprocessing, multimodal ML, calibration, XAI/SHAP, subgroup/fairness checks. Outputs: predicted pregnancy/live-birth risk, key drivers, personalized counseling/treatment suggestions, clinician review."),
    ("Candidate Topic 1", "Explainable Multimodal Machine Learning Framework for Personalized IVF Outcome Prediction and Clinical Decision Support in Indian Women."),
    ("Candidate Topic 2", "Trustworthy AI-Based Clinical Decision Support for Personalized Ovarian Stimulation and Embryo Transfer Outcome Prediction in IVF."),
    ("Candidate Topic 3", "Integration of Clinical, Embryological and Lifestyle Data for Explainable IVF Success Prediction and Patient-Centered Counseling."),
    ("Candidate Topic 4", "Explainable and Fair Machine Learning Models for IVF Live-Birth Prediction: External Validation and Indian Population Adaptation."),
    ("Current Supervisor Recommendation", "Do not finalize the title yet. The most defensible direction emerging from repeated gaps is explainable, personalized IVF decision support with external/Indian validation and multimodal clinical plus lifestyle data integration."),
]


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, max_width=55):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2, max_width)
        ws.column_dimensions[letter].width = max(width, 12)


def apply_body_style(ws):
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


wb = Workbook()
ws = wb.active
ws.title = "Literature Review Matrix"

headers = list(papers[0].keys())
ws.append(headers)
for p in papers:
    ws.append([p[h] for h in headers])
style_header(ws)
apply_body_style(ws)
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 84
auto_width(ws)
for col in ["B", "F", "G", "H", "K", "L", "M", "N"]:
    ws.column_dimensions[col].width = 42
ws.column_dimensions["U"].width = 55
tab = Table(displayName="LiteratureMatrix", ref=ws.dimensions)
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
ws.add_table(tab)

gap_counter = Counter()
gap_papers = defaultdict(list)
for p in papers:
    for gap in [g.strip() for g in p["Gap Found"].split(";")]:
        gap_counter[gap] += 1
        gap_papers[gap].append(str(p["Paper No"]))

ws2 = wb.create_sheet("Gap Frequency")
ws2.append(["Gap", "Paper Numbers", "Frequency", "Interpretation"])
interpretation = {
    "Need External Validation": "Most important repeated methodological gap; essential before novelty or clinical usefulness can be claimed.",
    "Need Clinical Decision Support": "Prediction needs translation into clinician-facing recommendation workflows.",
    "Retrospective Design": "Limits causal and deployment claims.",
    "Need Real-world Deployment Evidence": "Few models are tested in day-to-day IVF decision workflows.",
    "Limited Population Diversity": "Generalizability is uncertain across ethnic, clinical and geographic settings.",
    "No Indian Population": "Directly relevant to a Pune/India PhD context.",
    "Need Prospective/RCT Validation": "Required for AI-guided dose/trigger/transfer decision claims.",
}
for gap, freq in gap_counter.most_common():
    ws2.append([gap, ", ".join(gap_papers[gap]), freq, interpretation.get(gap, "Repeated or emerging gap; assess after additional papers are added.")])
style_header(ws2)
apply_body_style(ws2)
ws2.freeze_panes = "A2"
auto_width(ws2)
ws2.column_dimensions["D"].width = 70
tab2 = Table(displayName="GapFrequency", ref=ws2.dimensions)
tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
ws2.add_table(tab2)

chart = BarChart()
chart.title = "Top Repeated Gaps"
chart.y_axis.title = "Frequency"
chart.x_axis.title = "Gap"
data = Reference(ws2, min_col=3, min_row=1, max_row=min(ws2.max_row, 11))
cats = Reference(ws2, min_col=1, min_row=2, max_row=min(ws2.max_row, 11))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 8
chart.width = 18
ws2.add_chart(chart, "F2")

ws3 = wb.create_sheet("Synthesis")
ws3.append(["Item", "Supervisor-style synthesis"])
for item in deliverables:
    ws3.append(item)
style_header(ws3)
apply_body_style(ws3)
ws3.freeze_panes = "A2"
ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 105
for row in range(2, ws3.max_row + 1):
    ws3.row_dimensions[row].height = 48

ws4 = wb.create_sheet("Protocol Notes")
notes = [
    ("Scope", "AI/ML/XAI/CDSS/personalized IVF literature, emphasis 2021-2025. One 2026 systematic review is included only as a synthesis signal and flagged outside the target window."),
    ("Reading Strategy", "First-pass extraction uses abstract, discussion, conclusion, limitation and future-work sections where accessible. Full papers should be read only for shortlisted high-relevance rows."),
    ("Evidence Rule", "Do not treat every limitation as a gap. Candidate topic should be based on repeated gaps with methodological and population relevance."),
    ("Immediate Next Step", "Expand to 30-40 papers, verify Scopus/WoS indexing, read full limitations for the top 15, and add Indian IVF/ART data availability review."),
    ("Preferred Final Direction", "Explainable, personalized clinical decision support for IVF outcome prediction using multimodal clinical/embryology/lifestyle data, externally validated for Indian clinical settings."),
]
ws4.append(["Protocol Area", "Note"])
for note in notes:
    ws4.append(note)
style_header(ws4)
apply_body_style(ws4)
ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 110
for row in range(2, ws4.max_row + 1):
    ws4.row_dimensions[row].height = 45

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(OUT.resolve())
