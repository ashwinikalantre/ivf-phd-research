from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path("outputs/phd_ivf_lit_review/ivf_ai_literature_review_matrix_2021_2025.xlsx")
DOCS = Path("docs/research-design/themes")
DOCS.mkdir(parents=True, exist_ok=True)


def esc(value):
    return str(value or "").replace("|", "\\|").replace("\n", " ").strip()


def table(headers, rows):
    out = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        out.append("| " + " | ".join(esc(v) for v in row) + " |")
    return "\n".join(out)


wb = load_workbook(WORKBOOK, data_only=True)
ws = wb["Literature Review Matrix"]
headers = [c.value for c in ws[1]]
records = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
by_no = {r["Paper No"]: r for r in records}


themes = [
    {
        "slug": "ivf-outcome-prediction",
        "title": "IVF Outcome, Pregnancy and Live-Birth Prediction",
        "papers": [5, 6, 7, 8, 9, 10, 11, 15, 16, 17, 18, 19, 20, 29, 30, 31, 33],
        "existing_work": [
            "Most studies build predictive models for clinical pregnancy, implantation, miscarriage or live birth using routine clinical and treatment-cycle variables.",
            "Ensemble methods such as Random Forest, AdaBoost, XGBoost and voting classifiers frequently outperform simpler baselines, but several studies still show modest or context-dependent gains.",
            "Recent work is moving from general IVF prediction toward subgroup-specific prediction, including PCOS, endometriosis, male-factor infertility, FET and Indian subpopulation contexts.",
        ],
        "limitations": [
            "Many studies remain retrospective and single-center or geographically narrow.",
            "External validation is often missing or insufficient.",
            "Studies use different datasets, variables, outcomes and metrics, making direct comparison difficult.",
            "Prediction accuracy is often emphasized more than clinical workflow integration.",
        ],
        "usable_gaps": [
            "Externally validated IVF outcome prediction for Indian clinical settings.",
            "Standardized feature and reporting framework for IVF prediction studies.",
            "Explainable prediction outputs that can support counseling and treatment decisions.",
        ],
        "supervisor_view": "This theme is the broad evidence base. It supports the need for prediction, but alone it is too common for a PhD topic. The topic becomes stronger only when prediction is combined with explainability, clinical decision support and Indian/multimodal validation.",
    },
    {
        "slug": "embryo-selection-grading",
        "title": "Embryo Selection, Embryo Grading and Ploidy Prediction",
        "papers": [3, 4, 24, 26, 27, 28],
        "existing_work": [
            "Deep learning is widely applied to embryo images, time-lapse data, morphology and ploidy-related prediction.",
            "Some systems aim to automate embryo ranking and reduce inter-observer variation.",
            "The 2024 Nature Medicine randomized trial is important because it tests AI embryo selection prospectively rather than relying only on retrospective model performance.",
        ],
        "limitations": [
            "Image-only or transferred-embryo-only datasets may introduce selection bias.",
            "Black-box embryo selection raises ethical and trust concerns.",
            "The Nature Medicine RCT did not demonstrate noninferiority for clinical pregnancy, though it improved evaluation time.",
            "Many models are not trained or validated for the actual clinical decision: choosing among embryos of similar apparent quality.",
        ],
        "usable_gaps": [
            "Explainable embryo-selection support rather than opaque automated ranking.",
            "Clinical utility evaluation, not only AUC/accuracy.",
            "Integration of embryo features with clinical and patient context.",
        ],
        "supervisor_view": "Embryo AI is an important theme, but a thesis focused only on embryo image grading may become too laboratory-specific. It is better used as one component of a broader multimodal IVF decision-support framework.",
    },
    {
        "slug": "ovarian-stimulation-cdss",
        "title": "Ovarian Stimulation, Dose, Trigger and Clinical Decision Support",
        "papers": [1, 2, 21, 22, 23],
        "existing_work": [
            "AI and decision-support systems are being used to personalize ovarian stimulation, FSH dose, trigger timing and protocol selection.",
            "XAI follicle-size analysis shows how interpretable models can challenge rule-of-thumb clinical decisions.",
            "Real-world and randomized CDSS studies are beginning to appear, which is stronger evidence than retrospective prediction alone.",
        ],
        "limitations": [
            "Several CDSS studies are small, single-center or not randomized.",
            "Some systems combine multiple algorithms, making it difficult to isolate which component improves outcomes.",
            "Generalizability across clinics and physician behavior remains uncertain.",
            "Live-birth impact is not always directly measured.",
        ],
        "usable_gaps": [
            "Explainable CDSS for ovarian stimulation with external validation.",
            "Decision-support systems that show why dose/trigger suggestions are made.",
            "Evaluation of real-world usability and clinician trust.",
        ],
        "supervisor_view": "This is one of the strongest practical directions because it moves beyond prediction into action. It fits Computer Applications well if framed as an explainable CDSS rather than a medical protocol trial.",
    },
    {
        "slug": "fet-endometrium-ultrasound",
        "title": "FET, Endometrial Receptivity, Ultrasound and Radiomics",
        "papers": [10, 25, 33],
        "existing_work": [
            "Recent work uses ultrasound radiomics, endometrial features and clinical variables to predict FET outcomes, miscarriage or live birth.",
            "Fusion models combining imaging-derived scores with clinical data can outperform single-modality models.",
            "SHAP and other XAI tools are being used to provide individual-level explanation.",
        ],
        "limitations": [
            "Single-center datasets are common.",
            "Manual ROI drawing in ultrasound/radiomics may introduce subjectivity.",
            "FET-specific models may not generalize to fresh embryo transfer.",
            "External multicenter validation remains limited.",
        ],
        "usable_gaps": [
            "Multimodal endometrium plus clinical prediction with external validation.",
            "Automated or standardized imaging feature extraction.",
            "Explainable FET counseling support.",
        ],
        "supervisor_view": "This theme is promising but narrower. It can support the multimodal framework, especially if ultrasound/endometrial data are available. Without such data, it should remain a supporting theme.",
    },
    {
        "slug": "xai-trustworthy-ai",
        "title": "Explainable, Trustworthy and Patient-Centered AI",
        "papers": [1, 7, 12, 13, 14, 25, 26, 27, 32, 33, 34, 35],
        "existing_work": [
            "XAI methods such as SHAP, LIME and permutation importance are increasingly used in IVF/ART models.",
            "Ethical and patient-centered papers argue that IVF AI should be interpretable, contestable and understandable.",
            "Recent reviews emphasize validation, transparency, governance, fairness and clinical utility.",
        ],
        "limitations": [
            "Many studies use XAI as a technical add-on rather than evaluating whether clinicians or patients understand it.",
            "Patient-centered explainability remains underdeveloped.",
            "Trust, fairness and usability are rarely evaluated alongside model performance.",
            "Black-box embryo and treatment recommendations can create ethical risk.",
        ],
        "usable_gaps": [
            "Clinician-usable and patient-understandable explanation design for IVF predictions.",
            "Trustworthy AI evaluation framework including performance, calibration, XAI, subgroup fairness and usability.",
            "Shared-decision-support outputs rather than unexplained model scores.",
        ],
        "supervisor_view": "This is the intellectual core of the likely PhD topic. The contribution should not be another classifier; it should be an explainable, validated and usable decision-support framework.",
    },
    {
        "slug": "indian-lifestyle-data",
        "title": "Indian Population, Lifestyle Data and Contextual Personalization",
        "papers": [12, 14, 29, 35],
        "existing_work": [
            "Indian-population IVF AI evidence is limited but emerging.",
            "Personalization reviews highlight the need to adapt AI models to patient context and clinical setting.",
            "AI adoption surveys show growing clinical interest, but not necessarily enough outcome evidence.",
        ],
        "limitations": [
            "Only limited Indian IVF AI evidence was found in the current matrix.",
            "Lifestyle data are discussed as clinically relevant but are not consistently integrated in IVF AI models.",
            "Population-specific models need validation beyond one center.",
            "Socioeconomic and lifestyle factors may affect access, counseling and outcomes but are underrepresented in ML models.",
        ],
        "usable_gaps": [
            "Indian IVF validation dataset and population-specific model calibration.",
            "Integration of lifestyle and contextual factors with clinical/embryology variables.",
            "Personalized counseling framework adapted to Indian IVF patients.",
        ],
        "supervisor_view": "This theme can provide the local novelty. However, it must be strengthened carefully. We need more Indian/lifestyle evidence and a realistic data-access plan before making it the central title claim.",
    },
]


for theme in themes:
    rows = []
    for no in theme["papers"]:
        r = by_no.get(no)
        if not r:
            continue
        rows.append([
            r["Paper No"],
            r["Year"],
            r["Title"],
            r["Method/Algorithm"],
            r["Dataset"],
            r["Gap Found"],
        ])
    content = [
        f"# {theme['title']}",
        "",
        "## Papers in This Theme",
        "",
        table(["No", "Year", "Title", "Method", "Dataset", "Gap Found"], rows),
        "",
        "## Existing Work",
        "",
    ]
    content.extend([f"- {item}" for item in theme["existing_work"]])
    content.extend(["", "## Repeated Limitations", ""])
    content.extend([f"- {item}" for item in theme["limitations"]])
    content.extend(["", "## Usable Research Gaps", ""])
    content.extend([f"- {item}" for item in theme["usable_gaps"]])
    content.extend(["", "## Supervisor View", "", theme["supervisor_view"], ""])
    (DOCS / f"{theme['slug']}.md").write_text("\n".join(content), encoding="utf-8")


index_rows = [[theme["title"], f"[Open](themes/{theme['slug']}.md)", ", ".join(map(str, theme["papers"]))] for theme in themes]
Path("docs/research-design/theme-wise-review.md").write_text(
    "# Theme-Wise Review\n\n"
    "Phase 3 organizes the literature by research theme. Each theme page summarizes existing work, repeated limitations, usable gaps and supervisor interpretation.\n\n"
    + table(["Theme", "Page", "Paper Numbers"], index_rows)
    + "\n\n"
    "## Phase 3 Outcome\n\n"
    "The strongest theme combination is not simply IVF prediction. The stronger PhD direction is at the intersection of outcome prediction, explainability, clinical decision support, and Indian/contextual validation.\n",
    encoding="utf-8",
)

print("Exported theme review pages")
