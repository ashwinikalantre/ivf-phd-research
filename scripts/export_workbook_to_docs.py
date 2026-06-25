from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path("outputs/phd_ivf_lit_review/ivf_ai_literature_review_matrix_2021_2025.xlsx")
DOCS = Path("docs")


def md_escape(value):
    text = "" if value is None else str(value)
    return text.replace("|", "\\|").replace("\n", " ").strip()


def table(headers, rows):
    out = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        out.append("| " + " | ".join(md_escape(v) for v in row) + " |")
    return "\n".join(out)


wb = load_workbook(WORKBOOK, data_only=True)
ws = wb["Literature Review Matrix"]
headers = [c.value for c in ws[1]]
records = []
for row in ws.iter_rows(min_row=2, values_only=True):
    records.append(dict(zip(headers, row)))

gap_counter = Counter()
gap_papers = defaultdict(list)
for r in records:
    for gap in [g.strip() for g in str(r.get("Gap Found") or "").split(";") if g.strip()]:
        gap_counter[gap] += 1
        gap_papers[gap].append(r["Paper No"])

theme_counter = Counter()
theme_papers = defaultdict(list)
for r in records:
    for theme in [t.strip() for t in str(r.get("Category") or "").split(";") if t.strip()]:
        theme_counter[theme] += 1
        theme_papers[theme].append(r["Paper No"])

high_rows = []
for r in records:
    high_rows.append([
        r["Paper No"],
        r["Year"],
        r["Title"],
        r["Journal"],
        r["Objective"],
        r["Method/Algorithm"],
        r["Dataset"],
        r["Gap Found"],
        r["Relevance"],
    ])

(DOCS / "evidence-base/literature-matrix.md").write_text(
    "# Literature Matrix\n\n"
    "This page is a readable site view of the working Excel matrix. The Excel workbook remains the source for detailed extraction fields.\n\n"
    f"[Download the current Excel workbook](../assets/files/ivf_ai_literature_review_matrix_2021_2025.xlsx)\n\n"
    + table(
        ["No", "Year", "Title", "Journal", "Objective", "Method", "Dataset", "Gap Found", "Relevance"],
        high_rows,
    )
    + "\n",
    encoding="utf-8",
)

source_rows = []
for r in records:
    source_rows.append([
        r["Paper No"],
        r["Title"],
        r["Year"],
        r["Journal"],
        r["Source URL"],
        r["Extraction Confidence"],
    ])

(DOCS / "evidence-base/source-registry.md").write_text(
    "# Source Registry\n\n"
    "Use this page to audit where each paper came from and how reliable the current extraction is.\n\n"
    + table(["No", "Title", "Year", "Journal", "Source", "Extraction Confidence"], source_rows)
    + "\n\n"
    "## Cleaning Rules\n\n"
    "- Replace `Student document; URL pending` with DOI, PubMed, publisher, Scopus, or WoS links.\n"
    "- Verify all missing authors and journals before using a paper in the proposal.\n"
    "- Keep low-confidence papers as supporting/background evidence until full text is checked.\n",
    encoding="utf-8",
)

gap_rows = [[gap, ", ".join(map(str, gap_papers[gap])), freq] for gap, freq in gap_counter.most_common()]
(DOCS / "gap-analysis/gap-frequency.md").write_text(
    "# Gap Frequency\n\n"
    "This table is derived from the current literature matrix. It should guide topic discovery, but it should not be treated as final novelty proof.\n\n"
    + table(["Gap", "Paper Numbers", "Frequency"], gap_rows)
    + "\n",
    encoding="utf-8",
)

theme_rows = [[theme, ", ".join(map(str, theme_papers[theme])), freq] for theme, freq in theme_counter.most_common()]
(DOCS / "research-design/theme-classification.md").write_text(
    "# Theme Classification\n\n"
    "The current papers cluster around the following research themes.\n\n"
    + table(["Theme", "Paper Numbers", "Frequency"], theme_rows)
    + "\n\n"
    "## Working Theme Groups\n\n"
    "- IVF success, clinical pregnancy, and live-birth prediction\n"
    "- Embryo grading, embryo selection, and ploidy prediction\n"
    "- Ovarian stimulation, dose, trigger, and protocol decision support\n"
    "- FET, endometrial receptivity, ultrasound, and radiomics\n"
    "- Explainable AI, trustworthy AI, and patient-centered explanation\n"
    "- Multimodal data integration and clinical deployment\n",
    encoding="utf-8",
)

(DOCS / "evidence-base/excel-workbook.md").write_text(
    "# Excel Workbook\n\n"
    "The Excel workbook is the working evidence file for structured extraction and gap frequency analysis.\n\n"
    "[Download Excel workbook](../assets/files/ivf_ai_literature_review_matrix_2021_2025.xlsx)\n\n"
    "## Workbook Sheets\n\n"
    "- `Literature Review Matrix`: paper-level extraction.\n"
    "- `Gap Frequency`: repeated-gap count derived from paper categories.\n"
    "- `Synthesis`: research trend analysis, research problem, RQs, objectives, conceptual framework and candidate topics.\n"
    "- `Protocol Notes`: current rules and next steps.\n",
    encoding="utf-8",
)

print(f"Exported {len(records)} records into docs")
